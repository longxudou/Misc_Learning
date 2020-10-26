import re, json, os, argparse
import time, xlrd, xlsxwriter
from datetime import datetime
from collections import Counter
from Sequence_Classification.api import query, get_access_token
from openpyxl import load_workbook

TIME_INTERVAL = 30

access_token = get_access_token()


class item():
    def __init__(self, start_time, end_time, content):
        self.start_time = start_time
        self.end_time = end_time
        self.content = content

    def __repr__(self):
        # return f"{self.start_time.tm_hour}:{self.start_time.tm_min}:{self.start_time.tm_sec}-{self.end_time.tm_hour}:{self.end_time.tm_min}:{self.end_time.tm_sec}-{self.content}"

        return f"{self.start_time.tm_min}:{self.start_time.tm_sec}-{self.end_time.tm_min}:{self.end_time.tm_sec}-{self.content}"


def divide(path):
    with open(path, 'r') as f:
        lines = f.readlines()

    new_lines = []
    for i in lines:
        if i != '\n':
            new_lines.append(i.strip())
    lines = new_lines

    sent_cnt = len(lines) // 3

    item_list = []
    for index in range(sent_cnt):
        start = lines[index * 3 + 1][0:8]
        end = lines[index * 3 + 1][17:25]

        start = time.strptime(start, "%H:%M:%S")
        end = time.strptime(end, "%H:%M:%S")

        sentence = lines[index * 3 + 2]
        item_list.append(item(start, end, sentence))

    collect = []
    tmp = []
    for index in range(sent_cnt):
        cur_item = item_list[index]
        start = cur_item.start_time.tm_sec
        end = cur_item.end_time.tm_sec
        sentence = cur_item.content

        # 0, start, end , 30
        if 0 <= start <= end <= TIME_INTERVAL:
            # if tmp != [] and item_list[index-1].start_time.tm_min!=cur_item.start_time.tm_min:
            if tmp != [] and item_list[index - 1].start_time.tm_sec > TIME_INTERVAL:
                collect.append(tmp)
                tmp = [cur_item]
            else:
                tmp.append(cur_item)

        # 0, start, 30, end
        elif 0 <= start <= TIME_INTERVAL < end:
            tmp.append(cur_item)
            collect.append(tmp)
            tmp = []

        # 0, 30, start, 0, end
        elif TIME_INTERVAL <= start and 0 <= end < TIME_INTERVAL:
            tmp.append(cur_item)
            collect.append(tmp)
            tmp = []

        # 0, 30, start, end
        elif TIME_INTERVAL <= start <= end:
            if tmp != [] and item_list[index - 1].start_time.tm_sec < TIME_INTERVAL:
                collect.append(tmp)
                tmp = [cur_item]
            else:
                tmp.append(cur_item)

    if tmp != []:
        collect.append(tmp)
    return collect


def set_tag_for_interval(list_of_interval):
    result = []
    tag_list = []

    for list_of_sentence in list_of_interval:
        per_sentence_tag = query([i.content for i in list_of_sentence], access_token)
        tag_frequence = Counter([i[0] for i in per_sentence_tag])
        tag_list.append(tag_frequence.most_common(1)[0][0])

    cnt = 0
    for tag, interval in zip(tag_list, list_of_interval):
        interval_start = f"{interval[0].start_time.tm_hour}:{interval[0].start_time.tm_min}:{interval[0].start_time.tm_sec}"
        interval_end = f"{interval[-1].start_time.tm_hour}:{interval[-1].start_time.tm_min}:{interval[-1].start_time.tm_sec}"

        interval_content = ''.join([i.content for i in interval])
        result.append((f'{cnt}-{cnt + len(interval) - 1}', interval_start, interval_end, interval_content, tag))
        cnt = cnt + len(interval)

    return result


def write_xls(result, file_name):
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()

    # ('5-6', '0:0:31', '0:0:43', '啊。啊。', 'teach')
    # Write some data headers.
    worksheet.write('A1', 'Index')
    worksheet.write('B1', 'Start')
    worksheet.write('C1', 'End')
    worksheet.write('D1', 'Duration')
    worksheet.write('E1', 'Text')
    worksheet.write('G1', 'Tag')

    # Start from the first cell below the headers.
    row = 1
    col = 0

    for Index, Start, End, Duration, Text, Tag in (result):
        # Convert the date string into a datetime object.
        worksheet.write_string(row, col, Index)
        worksheet.write_string(row, col + 1, Start)
        worksheet.write_string(row, col + 2, End)
        worksheet.write_string(row, col + 3, Duration)
        worksheet.write_string(row, col + 4, Text)
        worksheet.write_string(row, col + 5, Tag)
        row += 1
    workbook.close()


def compute_duration(interval_list):
    result = []

    for interval in interval_list:
        start_time = list(map(int, interval[1].split(':')))
        end_time = list(map(int, interval[2].split(':')))

        # sec = (end_time[2] - start_time[2]) + (end_time[1] - start_time[1]) * 60 + (end_time[0] - start_time[0]) * 3600
        sec =(end_time[1] - start_time[1]) * 60 + (end_time[0] - start_time[0]) * 3600

        sec_start=0 if start_time[2]<=30 else 30
        sec_end=30 if start_time[2]<=30 else 60
        sec += sec_end-sec_start


        m = sec // 60
        s = sec % 60
        interval = list(interval)
        interval.insert(3, f"{m}:{s}")
        result.append(interval)

    return result


def merge_interval_based_on_tag(interval_list):
    result = []
    tmp = [interval_list[0]]
    for interval in interval_list[1:]:
        if interval[-1] == tmp[-1][-1]:
            tmp.append(interval)
        else:
            merge_interval = (f'{tmp[0][0].split("-")[0]}-{tmp[-1][0].split("-")[1]}',
                              tmp[0][1],
                              tmp[-1][2],
                              ''.join([i[3] for i in tmp]),
                              tmp[0][-1])
            result.append(merge_interval)
            tmp = [interval]
    if tmp != []:
        merge_interval = (f'{tmp[0][0].split("-")[0]}-{tmp[-1][0].split("-")[1]}',
                          tmp[0][1],
                          tmp[-1][2],
                          ''.join([i[3] for i in tmp]),
                          tmp[0][-1])
        result.append(merge_interval)
    return result


if __name__ == '__main__':
    # path_list=['/Users/longxud/Downloads/time/2020-03-10.srt']
    # for file_path in path_list:
    #     list_of_interval=divide(file_path)
    #     result=set_tag_for_interval(list_of_interval[:5])
    #     print(result)
    #     result=merge_interval_based_on_tag(result)
    #     result = compute_duration(result)
    #     write_xls(result,file_name='sample.xlsx')

    ###=======run on sample xlsx======
    path_list = ['2020-03-10.xlsx']
    for file_path in path_list:
        wb = load_workbook(filename=file_path)
        sheets = wb.get_sheet_names()
        sheet_first = sheets[0]
        ws = wb.get_sheet_by_name(sheet_first)
        rows = ws.rows
        columns = ws.columns

        test = []
        for row in rows:
            line = [col.value for col in row]
            test.append(line)

        test.pop(0)  # remove headers

        result = merge_interval_based_on_tag(test)
        result = compute_duration(result)
        write_xls(result, file_name='sample.xlsx')
